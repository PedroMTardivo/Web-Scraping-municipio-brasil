from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from openpyxl import load_workbook


wb = load_workbook("RUA_CEP.xlsx")
sheet = wb.active

driver = webdriver.Edge()
lista_de_link = []
driver.get("https://www.ruacep.com.br")
driver.maximize_window()
estados_LINK = driver.find_elements(By.XPATH,"//*[starts-with(@title, 'Estado') or starts-with(@title, 'Distrito')]")
for batata in estados_LINK:
    lista_de_link.append(batata.get_attribute('href'))
driver.switch_to.new_window('tab')
for sla in lista_de_link:
    
    driver.get(sla)
    

    tnt =1
    count = 0 
    while tnt == 1:
        
        cidades = driver.find_elements(By.XPATH,'//div [@class="row"]//div[@class = "col-sm-6 mb-4"]//a//strong')
        for cidade in cidades:
            nome = cidade.text
            nome = nome.split(", ")
            nome_cidade = nome[0]
            estado = nome[1]
            
            linhaf = sheet.max_row
            sheet.cell(row=linhaf+1,column=1,value=estado)
            sheet.cell(row=linhaf+1,column=2,value=nome_cidade)
            
            ceps = driver.find_elements(By.XPATH,'//div [@class="row"]//div[@class = "col-sm-6 mb-4"]//*[contains(text(),"CEP")]')
            for cep in ceps:
                cep_n = cep.text
                cep_n = cep_n.split("\n")[0]
                cep_n = cep_n.split("CEP: ")[-1]
                sheet.cell(row=linhaf+1,column=3,value=cep_n)
        sleep(2)
        pagina = driver.find_elements(By.XPATH,'//*//li[@class="page-item"]//a[@class="page-link"]')
        if len(pagina) <=1:
            count+=1
            if count ==2:
                count= 0 
                
                break
        try:    
            pagina = pagina[-1]
        except:
            continue
        pagina = pagina.get_attribute('href')
        driver.get(pagina)

wb.save("RUA_CEP_FINAL.xlsx")
